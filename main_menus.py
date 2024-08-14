from Idées_menus import *
from fonctions import *
import openpyxl
import random
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from configparser import ConfigParser


config = ConfigParser()
config.read('C:\\Users\\Sandra\\Desktop\\Programme Menus\\config\\keys_config.cfg')

MAIL_KEY = config.get('Mailexp', 'user_key')
MAIL_USER = config.get('Mailexp', 'mail_user')
MAIL_DEST = config.get('Maildest', 'maildest_user')


liste = random.sample(saison(), k=14)

menus_dict = {index: menus for index, menus in enumerate(liste)}

#Ouverture du classeur Excel

wb = openpyxl.load_workbook('MENUS 2024.xlsx')
ws = wb['MENUS']

#Ajout des menus à chaque cellule désignée

ws.cell(5,2,menus_dict.get(0))
ws.cell(5,3,menus_dict.get(1))
ws.cell(5,4,menus_dict.get(2))
ws.cell(5,5,menus_dict.get(3))
ws.cell(5,6,menus_dict.get(4))
ws.cell(5,7,menus_dict.get(5))
ws.cell(5,8,menus_dict.get(6))
ws.cell(7,2,menus_dict.get(7))
ws.cell(7,3,menus_dict.get(8))
ws.cell(7,4,menus_dict.get(9))
ws.cell(7,5,menus_dict.get(10))
ws.cell(7,6,menus_dict.get(11))
ws.cell(7,7,menus_dict.get(12))
ws.cell(7,8,menus_dict.get(13))


wb.save('MENUS 2024.xlsx')


#Informations de connexion au serveur SMTP avec l'adresse du serveur concerné (Outlook, gmail etc)

smtp_server = 'smtp-mail.outlook.com'
port = 587
sender_email = MAIL_USER
password = MAIL_KEY
receiver_email = MAIL_DEST  


#Création du mail

message = MIMEMultipart()
message['From'] = sender_email
message['To'] = receiver_email
message['Subject'] = 'Menus de la semaine'
body = ''
message.attach(MIMEText(body, 'plain'))



#Ajout de la pièce jointe

filename = 'MENUS 2024.xlsx'
attachment = open(filename, "rb")
part = MIMEBase('application', 'octet-stream')
part.set_payload(attachment.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
message.attach(part)

#Envoi du message en se connectant au serveur via les logs fournits

with smtplib.SMTP(smtp_server, port) as server:
    server.starttls()
    server.login(sender_email, password)
    text = message.as_string()
    server.sendmail(sender_email, receiver_email, text)
